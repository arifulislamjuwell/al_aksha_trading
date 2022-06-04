import datetime
from io import BytesIO
import openpyxl
from django.db import connection
from django.http import HttpResponse
from utils.transaction import my_transasction_dict_make, customer_transasction_dict_make
from dashboard.models import MyTransaction


def last_day_of_month(any_month):
    next_month = any_month.replace(
        day=28) + datetime.timedelta(days=4)  # this will never fail
    return next_month - datetime.timedelta(days=next_month.day)


def my_transaction_download(transactions):
    final_data = my_transasction_dict_make(transactions, 0)
    sheet_name = 'My Transaction'
    starting_row = 2
    output = BytesIO()

    workbook = openpyxl.load_workbook(filename='sample/my_transaction_sample.xlsx')
    sheet = workbook.worksheets[0]

    for record in final_data:
        value_index = 0
        for index, value in enumerate(['date', 'transaction_type', 'quantity', 'details', 'total_bill', 'paid', 'current_balance']):
            sheet.cell(row=starting_row,column=index + 1).value = record.get(value)
            value_index += 1
        starting_row += 1

    workbook.save(output)
    output.seek(0)

    filename = str(sheet_name) + ".xlsx"
    content_disposition = "attachment; filename={}".format(filename)
    response = HttpResponse(output.read(
    ), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = content_disposition

    return response


def customer_transaction_download(transactions):
    final_data = customer_transasction_dict_make(transactions, 0)
    sheet_name = 'Customer Transaction'
    starting_row = 2
    output = BytesIO()

    workbook = openpyxl.load_workbook(filename='sample/my_transaction_sample.xlsx')
    sheet = workbook.worksheets[0]

    for record in final_data:
        value_index = 0
        for index, value in enumerate(['date', 'transaction_type', 'quantity', 'details', 'total_bill', 'paid', 'current_balance']):
            sheet.cell(row=starting_row,column=index + 1).value = record.get(value)
            value_index += 1
        starting_row += 1

    workbook.save(output)
    output.seek(0)

    filename = str(sheet_name) + ".xlsx"
    content_disposition = "attachment; filename={}".format(filename)
    response = HttpResponse(output.read(
    ), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = content_disposition

    return response